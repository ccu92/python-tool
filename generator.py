
import openpyxl, time
wb = openpyxl.load_workbook('出入庫管理.xlsx')
sheet = wb['工作表1']
maxRow = sheet.max_row
startFromHere = maxRow

while True:    
    isValueNone = sheet.cell(row = startFromHere, column = 6).value
    if isValueNone is None != True:
        startFromHere = startFromHere - 1          
    else:        
        break
        

       
#已經檢查到有值的格子

finalCell = 'L' + str(startFromHere)

def ListAll():
    tuple(sheet['A1': finalCell])
    myList = []
    for rowOfCellObjects in sheet['A1': finalCell]:
        
        for cellObj in rowOfCellObjects:
            myList.append((cellObj.coordinate, cellObj.value))
    return myList
            
listAll = ListAll()        
lastRow = listAll[len(listAll)-12: len(listAll)]

print('     ***出貨資料***')
print('機尾號:     ' + lastRow[1][1])
print('------------------------------')
print('Part No.:   ' + lastRow[2][1])
print('------------------------------')
print('Serial No.: ' + lastRow[3][1])
print('------------------------------')
print('描述:       ' + lastRow[4][1])
print('------------------------------')
print('PO No.:     ' + lastRow[5][1])
print('------------------------------')
print('數量:       ' + str(lastRow[7][1]))
print('------------------------------')
print('測試版，請按任意鍵退出')
print('Beta version, please press any key to exit.')




wb2 = openpyxl.load_workbook('template.xlsx')
sheet2 = wb2['Sheet']


sheet2['B7'] = lastRow[2][1]
sheet2['C7'] = lastRow[4][1]
sheet2['D7'] = lastRow[5][1]
sheet2['E7'] = lastRow[3][1]
sheet2['G7'] = lastRow[7][1]
sheet2['F13'] = lastRow[1][1]


def UserName():
    if lastRow[1][1] == 'B-99888':
        return '鋐維'
    elif lastRow[1][1] == 'B-99988':
        return 'TIGER HERCULES'
    else:
        return '寶豐隆'
    


userName = UserName()
sheet2['F4'] = userName

def SheetSerialNumber():
    num = lastRow[0][1]
    sNum = num[-4:-1]
    yyyyMmDd = lastRow[11][1]
    serialNum = yyyyMmDd.strftime("%Y%m%d") + sNum
    return serialNum

sheetSerialNumber = SheetSerialNumber()
sheet2['B2'] = sheetSerialNumber

def TimeStamp():
    yyyyMmDd = lastRow[11][1]
    timeStamp = yyyyMmDd.strftime('%Y%m%d')
    timeStamp = timeStamp[0:4] + '年' + timeStamp[4:6] + '月' +  timeStamp[6:8] + '日'
    return timeStamp


timeStamp = TimeStamp()
sheet2['F2'] = timeStamp





#wb2.save('draft.xlsx')
wb2.save(lastRow[1][1] + ' ' + lastRow[4][1] + ' ' + sheetSerialNumber + '.xlsx')
input()
